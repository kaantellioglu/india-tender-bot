"""
Sonuclari, mevcut 'India_Procurement_Intelligence_Database.xlsx' ile AYNI
sema/sekmelerle guncellenen bir calisma dosyasina yazar.

v2 iyilestirmeleri:
- Daha once eklenmis URL'lerde bos alanlar varsa yeni taramada bulunan bilgilerle
  zenginlestirilir. Boylece dashboard'daki eski kayitlar da zamanla dolar.
- lead.extra icindeki tender_ref, tender_date, closing_date, source_type, score
  bilgileri Notes/ilgili kolonlara yazilir.
- Status LOA/AOC/FOA/award sinyallerine gore "Awarded" olabilir.
"""
from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl

from ..scrapers.base_scraper import TenderLead
from ..parsers.pdf_parser import ExtractedTenderInfo

logger = logging.getLogger(__name__)

TENDER_REGISTER_SHEET = "04 Tender Register"
PRICE_INTEL_SHEET = "06 Price Intelligence"
SOURCE_LOG_SHEET = "08 Source Log"

TENDER_REGISTER_HEADER_ROW = 3
PRICE_INTEL_HEADER_ROW = 3

# 04 Tender Register kolon indeksleri (1-indexed)
COL = {
    "tender_id": 1,
    "company": 2,
    "portal": 3,
    "tender_ref": 4,
    "tender_date": 5,
    "status": 6,
    "product_segment": 7,
    "product_description": 8,
    "qty": 9,
    "uom": 10,
    "winner": 11,
    "brand": 12,
    "total_inr_basic": 13,
    "gst_inr": 14,
    "total_inr_incl_gst": 15,
    "fx_inr_eur": 16,
    "total_eur": 17,
    "unit_eur_basic": 18,
    "unit_eur_incl_gst": 19,
    "source_url": 20,
    "pdf_saved": 21,
    "notes": 22,
}


def _is_blank(value) -> bool:
    return value is None or str(value).strip() == "" or str(value).strip().lower() in {"unknown", "not found", "—"}


def _clean(value, limit: int | None = None):
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text[:limit] if limit else text


def _existing_url_rows(ws, url_col_idx: int, header_row: int) -> dict[str, int]:
    urls: dict[str, int] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row_idx, url_col_idx).value
        if val:
            urls[str(val).strip()] = row_idx
    return urls


def _next_tender_id(ws, id_col_idx: int, header_row: int, prefix: str = "T-") -> str:
    max_num = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_col=id_col_idx, values_only=True):
        val = row[id_col_idx - 1]
        if val and str(val).startswith(prefix):
            try:
                num = int(str(val).replace(prefix, ""))
                max_num = max(max_num, num)
            except ValueError:
                continue
    return f"{prefix}{max_num + 1:04d}"


def _infer_status(lead: TenderLead, info: ExtractedTenderInfo | None) -> str:
    hay = f"{lead.title} {lead.url} {lead.raw_snippet or ''}".lower()
    if lead.extra.get("awarded_candidate") or any(x in hay for x in ["aoc", "loa", "foa", "award", "successful bidder"]):
        return "Awarded"
    if info and (info.tender_ref or info.text_excerpt):
        return "Docs Found"
    return "New Lead"


def _product_segment(lead: TenderLead, info: ExtractedTenderInfo | None) -> str:
    return (
        lead.matched_keyword
        or (info.product_segment if info else None)
        or ""
    )


def _note_for_lead(lead: TenderLead, info: ExtractedTenderInfo | None, enriched: bool = False) -> str:
    parts = [
        "Bot enrich" if enriched else "Bot tarafindan tespit edildi",
        f"date={datetime.now():%Y-%m-%d}",
    ]
    if lead.file_type:
        parts.append(f"type={lead.file_type}")
    if lead.extra.get("closing_date"):
        parts.append(f"closing={lead.extra.get('closing_date')}")
    if info and info.closing_date:
        parts.append(f"pdf_closing={info.closing_date}")
    if lead.extra.get("score") is not None:
        parts.append(f"score={lead.extra.get('score')}")
    if info and info.confidence:
        parts.append(f"pdf_conf={info.confidence}")
    return "; ".join(parts)


def _set_if_blank(ws, row_idx: int, col_idx: int, value) -> bool:
    if value is None or value == "":
        return False
    cell = ws.cell(row_idx, col_idx)
    if _is_blank(cell.value):
        cell.value = value
        return True
    return False


def _append_note(ws, row_idx: int, note: str) -> None:
    cell = ws.cell(row_idx, COL["notes"])
    existing = _clean(cell.value)
    if existing and note not in existing:
        cell.value = f"{existing} | {note}"
    elif not existing:
        cell.value = note


def update_workbook(
    workbook_path: str | Path,
    leads: Iterable[TenderLead],
    extracted_by_url: dict[str, ExtractedTenderInfo],
) -> dict:
    workbook_path = Path(workbook_path)
    wb = openpyxl.load_workbook(workbook_path)

    ws_reg = wb[TENDER_REGISTER_SHEET]
    ws_price = wb[PRICE_INTEL_SHEET]
    ws_log = wb[SOURCE_LOG_SHEET]

    existing_url_rows = _existing_url_rows(ws_reg, COL["source_url"], TENDER_REGISTER_HEADER_ROW)

    new_register_rows = 0
    new_price_rows = 0
    enriched_rows = 0

    for lead in leads:
        info = extracted_by_url.get(lead.url)

        lead_tender_ref = lead.extra.get("tender_ref")
        lead_tender_date = lead.extra.get("tender_date") or lead.published_date
        tender_ref = (info.tender_ref if info and info.tender_ref else lead_tender_ref)
        tender_date = (info.tender_date if info and info.tender_date else lead_tender_date)
        closing_date = (info.closing_date if info and info.closing_date else lead.extra.get("closing_date"))

        if lead.url in existing_url_rows:
            row_idx = existing_url_rows[lead.url]
            changed = False
            changed |= _set_if_blank(ws_reg, row_idx, COL["tender_ref"], tender_ref)
            changed |= _set_if_blank(ws_reg, row_idx, COL["tender_date"], tender_date)
            changed |= _set_if_blank(ws_reg, row_idx, COL["product_segment"], _product_segment(lead, info))
            changed |= _set_if_blank(ws_reg, row_idx, COL["product_description"], _clean(lead.title, 500))
            changed |= _set_if_blank(ws_reg, row_idx, COL["qty"], info.qty if info else None)
            changed |= _set_if_blank(ws_reg, row_idx, COL["uom"], info.unit if info else None)
            changed |= _set_if_blank(ws_reg, row_idx, COL["winner"], info.winner if info else None)
            changed |= _set_if_blank(ws_reg, row_idx, COL["total_inr_basic"], info.total_price_inr if info else None)
            changed |= _set_if_blank(ws_reg, row_idx, COL["pdf_saved"], "Yes" if (info and info.text_excerpt) else None)

            # New Lead ise ama AOC/LOA sinyali geldiyse status'u yukseltebiliriz.
            current_status = ws_reg.cell(row_idx, COL["status"]).value
            new_status = _infer_status(lead, info)
            if new_status == "Awarded" and current_status != "Awarded":
                ws_reg.cell(row_idx, COL["status"]).value = "Awarded"
                changed = True

            if closing_date:
                _append_note(ws_reg, row_idx, f"closing={closing_date}")
                changed = True

            if changed:
                _append_note(ws_reg, row_idx, _note_for_lead(lead, info, enriched=True))
                enriched_rows += 1
            continue

        tender_id = _next_tender_id(ws_reg, COL["tender_id"], TENDER_REGISTER_HEADER_ROW)
        status = _infer_status(lead, info)

        row_values = [
            tender_id,
            lead.portal_name,
            lead.portal_name,
            tender_ref,
            tender_date,
            status,
            _product_segment(lead, info),
            _clean(lead.title, 500),
            info.qty if info else None,
            info.unit if info else None,
            info.winner if info else "Not found",
            "Unknown",
            info.total_price_inr if info else None,
            None,
            None,
            None,
            None,
            None,
            None,
            lead.url,
            "Yes" if (info and info.text_excerpt) else "No",
            _note_for_lead(lead, info),
        ]

        ws_reg.append(row_values)
        existing_url_rows[lead.url] = ws_reg.max_row
        new_register_rows += 1

        # Fiyat/miktar cikarilabildiyse Price Intelligence'a da ekle.
        if info and (info.total_price_inr or info.unit_price_inr or info.qty):
            price_row = [
                tender_id,
                lead.portal_name,
                datetime.now().year,
                _product_segment(lead, info),
                None,
                info.qty,
                info.winner or "Not found",
                "Unknown",
                info.total_price_inr,
                None,
                info.unit_price_inr,
                None,
                info.confidence,
                lead.url,
                "Bot tarafindan tespit edildi, dogrulama gerekli",
            ]
            ws_price.append(price_row)
            new_price_rows += 1

    ws_log.append([
        f"{datetime.now():%Y-%m-%d %H:%M}",
        f"Otomatik tarama: {new_register_rows} yeni ihale, {new_price_rows} yeni fiyat kaydi, {enriched_rows} mevcut satir zenginlestirildi",
        "bot",
    ])

    wb.save(workbook_path)
    logger.info(
        "Workbook guncellendi: %s yeni ihale, %s yeni fiyat kaydi, %s enrich",
        new_register_rows, new_price_rows, enriched_rows,
    )

    return {
        "new_register_rows": new_register_rows,
        "new_price_rows": new_price_rows,
        "enriched_rows": enriched_rows,
    }
