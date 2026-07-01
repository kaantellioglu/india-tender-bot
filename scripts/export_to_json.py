"""Export clean market-intelligence dataset to docs/data.json."""
from __future__ import annotations
import argparse, json, re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
import openpyxl, yaml
from src.quality.market_filter import classify_market_record
ROOT=Path(__file__).resolve().parent.parent; DEFAULT_WORKBOOK=ROOT/'data'/'India_Procurement_Intelligence_Database.xlsx'; DEFAULT_OUT=ROOT/'docs'/'data.json'; FAILURES_PATH=ROOT/'data'/'source_failures.json'; ACTIONS_PATH=ROOT/'data'/'action_queue.json'; SOURCE_CLEANUP_PATH=ROOT/'config'/'source_cleanup.yaml'

def rows_as_dicts(ws, header_row, headers):
    out=[]
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, values_only=True):
        if row is None or all(v is None for v in row) or row[0] is None: continue
        rec={}
        for h,v in zip(headers,row):
            if isinstance(v, datetime): v=v.strftime('%Y-%m-%d')
            rec[h]=v
        out.append(rec)
    return out

def load_json_list(path):
    if not path.exists(): return []
    try:
        data=json.loads(path.read_text(encoding='utf-8')); return data if isinstance(data,list) else []
    except Exception: return []

def load_source_cleanup(path=SOURCE_CLEANUP_PATH):
    if not path.exists(): return []
    try:
        data=yaml.safe_load(path.read_text(encoding='utf-8')) or {}; rows=[]
        for portal_id,item in (data.get('sources') or {}).items():
            if isinstance(item,dict):
                row={'portal_id':portal_id,**item}; row.setdefault('scope','market_intelligence_only'); row.setdefault('excluded_scope','no_procurement_transaction_automation'); rows.append(row)
        return rows
    except Exception: return []

def parse_meta(notes):
    notes=str(notes or ''); meta={}; m=re.search(r'meta:(.*)',notes)
    if not m: return meta
    for part in m.group(1).split(';'):
        if '=' in part:
            k,v=part.split('=',1); meta[k.strip()]=v.strip()
    return meta

def safe_int(v,default=0):
    try: return default if v in (None,'') else int(float(str(v)))
    except Exception: return default

def safe_float(v):
    try: return None if v in (None,'') else float(str(v).replace(',',''))
    except Exception: return None

def derive_tender_fields(t):
    meta=parse_meta(t.get('notes'))
    hay=' '.join(str(x or '') for x in [t.get('product_segment'),t.get('product_description'),t.get('tender_ref'),t.get('status'),t.get('source_url'),meta.get('equipment'),meta.get('matched_terms')])
    q=classify_market_record(hay, source_url=str(t.get('source_url') or ''))
    t['lead_score']=safe_int(meta.get('score'),0); t['priority']=meta.get('priority') or ('High' if t['lead_score']>=70 else 'Medium' if t['lead_score']>=40 else 'Low')
    t['source_type']=meta.get('source_type') or ('pdf' if str(t.get('source_url') or '').lower().split('?')[0].endswith('.pdf') else 'html')
    t['closing_date']=meta.get('closing_date') or ''; t['equipment']=meta.get('equipment') or q.equipment_segment or t.get('product_segment') or ''
    t['quality_status']=meta.get('quality_status') or q.quality_status; t['market_relevance']=meta.get('market_relevance') or q.market_relevance; t['document_type']=meta.get('document_type') or q.document_type
    t['unit_price_inr']=safe_float(meta.get('unit_price_inr')) or safe_float(t.get('unit_eur_basic')); t['amount_label']=meta.get('amount_label') or ''
    t['extraction_scope']=meta.get('scope') or 'market_intelligence_only'; t['excluded_scope']=meta.get('excluded_scope') or 'no_bid_or_offer_submission_automation'
    t['missing_ref']=not bool(t.get('tender_ref')); t['missing_date']=not bool(t.get('tender_date')); t['missing_qty']=not bool(t.get('qty')); t['missing_winner']=not bool(t.get('winner')) or str(t.get('winner')).lower() in {'not found','unknown','none'}; t['missing_price']=not bool(t.get('total_inr_basic')) and not bool(t.get('unit_price_inr'))
    t['is_clean_market_target']=t['quality_status']=='market_target'; t['needs_quality_review']=t['quality_status']=='needs_review'
    return t

def build_portal_health(portals,tenders,failures,actions,cleanup=None):
    by_portal_tenders=Counter(t.get('company') or t.get('portal') or '' for t in tenders); by_portal_failures=defaultdict(list); by_portal_actions=defaultdict(list); cleanup=cleanup or []; cleanup_by_name={c.get('portal'):c for c in cleanup}; cleanup_by_id={c.get('portal_id'):c for c in cleanup}
    for f in failures: by_portal_failures[f.get('portal') or f.get('portal_id')].append(f)
    for a in actions: by_portal_actions[a.get('portal') or a.get('portal_id')].append(a)
    health=[]
    for p in portals:
        name=p.get('name') or ''; fl=by_portal_failures.get(name,[])+by_portal_failures.get(p.get('id'),[]); ac=by_portal_actions.get(name,[])+by_portal_actions.get(p.get('id'),[]); cu=cleanup_by_name.get(name) or cleanup_by_id.get(p.get('id')) or {}; status='OK'
        if any((f.get('failure_type') or '').startswith('http_403') for f in fl) or ac: status='Access Review'
        elif any(f.get('failure_type') in {'dns_error','ssl_error','timeout','connection_refused','redirect_loop'} for f in fl): status='Access Problem'
        elif any((f.get('failure_type') or '').startswith('http_404') for f in fl): status='Broken URL'
        elif fl: status='Review'
        if cu.get('status') in {'source_repair_required','broken_domain','access_problem','protected_or_rate_limited'}: status='Source Cleanup'
        elif cu.get('status') in {'priority_source','priority_aggregator','improved_source'} and status=='OK': status='Priority Source'
        health.append({'portal':name,'portal_id':p.get('id',''),'tier':p.get('tier',''),'relevance':p.get('relevance',''),'status':status,'lead_count':by_portal_tenders.get(name,0),'failure_count':len(fl),'open_access_review_count':len([x for x in ac if x.get('status','open')=='open']),'top_failure':Counter(f.get('failure_type') for f in fl).most_common(1)[0][0] if fl else '','next_action':cu.get('technical_fix_needed') or (ac[-1].get('next_action') if ac else p.get('next_action','')),'data_access_status':cu.get('data_access_status',''),'market_intelligence_priority':cu.get('market_intelligence_priority',''),'public_tender_url':cu.get('public_tender_url',''),'archive_award_url':cu.get('archive_award_url',''),'website':p.get('website','')})
    return health

def clean_price_rows(rows):
    out=[]
    for r in rows:
        q=classify_market_record(' '.join(str(r.get(k) or '') for k in ['product_segment','notes','source_url','company']), source_url=str(r.get('source_url') or ''))
        if q.quality_status=='market_target': r['quality_status']=q.quality_status; r['equipment']=q.equipment_segment; out.append(r)
    return out

def export(workbook_path,out_path):
    wb=openpyxl.load_workbook(workbook_path,data_only=True)
    portals=rows_as_dicts(wb['01 Portal Master'],3,['tier','name','type','website','tender_search_url','award_url','login_requirement','relevance','note','next_action'])
    raw=[derive_tender_fields(t) for t in rows_as_dicts(wb['04 Tender Register'],3,['tender_id','company','portal','tender_ref','tender_date','status','product_segment','product_description','qty','uom','winner','brand','total_inr_basic','gst_inr','total_inr_incl_gst','fx_inr_eur','total_eur','unit_eur_basic','unit_eur_incl_gst','source_url','pdf_saved','notes'])]
    clean=[t for t in raw if t.get('is_clean_market_target')]; review=[t for t in raw if t.get('needs_quality_review')]; rejected=[t for t in raw if t.get('quality_status')=='non_target']
    price=clean_price_rows(rows_as_dicts(wb['06 Price Intelligence'],3,['tender_id','company','year','product_segment','capacity','qty','winner','brand','total_inr_basic','total_eur_basic','unit_eur_basic','unit_eur_incl_gst','confidence','source_url','notes']))
    comps=rows_as_dicts(wb['07 Competitor DB'],3,['name','country','role','segment','website','relevance','notes']); log=rows_as_dicts(wb['08 Source Log'],3,['timestamp','message','actor'])[-100:]
    failures=load_json_list(FAILURES_PATH); actions=load_json_list(ACTIONS_PATH); cleanup=load_source_cleanup(); health=build_portal_health(portals,clean,failures,actions,cleanup)
    dashboard={'generated_at':datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC'),'portal_count':len(portals),'raw_tender_count':len(raw),'tender_count':len(clean),'review_tender_count':len(review),'rejected_tender_count':len(rejected),'price_row_count':len(price),'competitor_count':len(comps),'awarded_count':sum(1 for t in clean if 'award' in str(t.get('status') or '').lower()),'high_priority_count':sum(1 for t in clean if t.get('priority')=='High'),'medium_priority_count':sum(1 for t in clean if t.get('priority')=='Medium'),'failure_count':len(failures),'open_access_review_count':len([a for a in actions if a.get('status','open')=='open']),'portal_access_review_count':sum(1 for p in health if p.get('status') in {'Access Review','Access Problem','Broken URL','Source Cleanup'}),'source_cleanup_count':len(cleanup),'missing_ref_count':sum(1 for t in clean if t.get('missing_ref')),'missing_date_count':sum(1 for t in clean if t.get('missing_date')),'missing_qty_count':sum(1 for t in clean if t.get('missing_qty')),'missing_price_count':sum(1 for t in clean if t.get('missing_price')),'scope':'market_intelligence_only','excluded_scope':'no_bid_or_offer_submission_automation'}
    payload={'dashboard':dashboard,'portals':portals,'tenders':clean,'raw_tenders':raw[-2000:],'quality_review':review[-1000:],'rejected_tenders':rejected[-1000:],'price_intelligence':price,'competitors':comps,'source_log':log,'source_failures':failures[-1000:],'action_queue':actions[-1000:],'access_queue':actions[-1000:],'portal_health':health,'source_cleanup':cleanup}
    out_path.parent.mkdir(parents=True,exist_ok=True); text=json.dumps(payload,ensure_ascii=False,indent=2)
    if len(text)<100: raise RuntimeError('Export produced suspiciously small data.json')
    out_path.write_text(text,encoding='utf-8'); print(f'Yazildi: {out_path} ({len(portals)} portal, {len(clean)} clean ihale, {len(review)} review, {len(price)} fiyat)')

def parse_args():
    p=argparse.ArgumentParser(); p.add_argument('--workbook',default=str(DEFAULT_WORKBOOK)); p.add_argument('--out',default=str(DEFAULT_OUT)); return p.parse_args()
if __name__=='__main__': args=parse_args(); export(Path(args.workbook),Path(args.out))
